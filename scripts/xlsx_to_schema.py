#!/usr/bin/env python3
"""Convert a metadata Excel file to FlowScope SchemaMetadata JSON.

The expected xlsx format has columns: SCHEMA, TABLE_NAME, COLUMN_NAME
(optionally DATA_TYPE). Each row describes one column of one table.

Usage:
    python scripts/xlsx_to_schema.py metadata.xlsx -o oracle_schema.json
    python scripts/xlsx_to_schema.py metadata.xlsx --upper  # force uppercase

Requires: openpyxl  (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import OrderedDict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def read_xlsx(path: str) -> list[dict]:
    """Read an xlsx file and return rows as dicts keyed by header names."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().upper() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        entry = {}
        for header, val in zip(headers, row):
            if header:
                entry[header] = str(val).strip() if val is not None else ""
        result.append(entry)
    wb.close()
    return result


def rows_to_schema(
    rows: list[dict],
    upper: bool = False,
    schema_col: str = "SCHEMA",
    table_col: str = "TABLE_NAME",
    column_col: str = "COLUMN_NAME",
    type_col: str = "DATA_TYPE",
) -> dict:
    """Convert flat rows into FlowScope SchemaMetadata JSON structure."""
    # Group by (schema, table)
    tables: OrderedDict[tuple[str, str], list[dict]] = OrderedDict()
    for row in rows:
        schema_name = row.get(schema_col, "")
        table_name = row.get(table_col, "")
        col_name = row.get(column_col, "")
        data_type = row.get(type_col, "")

        if not table_name or not col_name:
            continue

        if upper:
            schema_name = schema_name.upper()
            table_name = table_name.upper()
            col_name = col_name.upper()

        key = (schema_name, table_name)
        col_entry = {"name": col_name}
        if data_type:
            col_entry["dataType"] = data_type.upper() if upper else data_type

        tables.setdefault(key, []).append(col_entry)

    schema_tables = []
    for (schema_name, table_name), columns in tables.items():
        entry = {"name": table_name, "columns": columns}
        if schema_name:
            entry["schema"] = schema_name
        # Insert schema before name for readability
        ordered = {}
        if "schema" in entry:
            ordered["schema"] = entry["schema"]
        ordered["name"] = entry["name"]
        ordered["columns"] = entry["columns"]
        schema_tables.append(ordered)

    return {
        "defaultCatalog": None,
        "defaultSchema": None,
        "caseSensitivity": "Upper" if upper else None,
        "allowImplied": True,
        "tables": schema_tables,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert metadata xlsx to FlowScope SchemaMetadata JSON"
    )
    parser.add_argument("xlsx", help="Path to the metadata Excel file")
    parser.add_argument("-o", "--output", help="Output JSON file (default: stdout)")
    parser.add_argument(
        "--upper",
        action="store_true",
        help="Force all identifiers to UPPERCASE (Oracle convention)",
    )
    parser.add_argument(
        "--schema-col",
        default="SCHEMA",
        help="Header name for schema column (default: SCHEMA)",
    )
    parser.add_argument(
        "--table-col",
        default="TABLE_NAME",
        help="Header name for table column (default: TABLE_NAME)",
    )
    parser.add_argument(
        "--column-col",
        default="COLUMN_NAME",
        help="Header name for column column (default: COLUMN_NAME)",
    )
    parser.add_argument(
        "--type-col",
        default="DATA_TYPE",
        help="Header name for data type column (default: DATA_TYPE)",
    )
    args = parser.parse_args()

    rows = read_xlsx(args.xlsx)
    if not rows:
        sys.exit(f"No data rows found in {args.xlsx}")

    schema = rows_to_schema(
        rows,
        upper=args.upper,
        schema_col=args.schema_col,
        table_col=args.table_col,
        column_col=args.column_col,
        type_col=args.type_col,
    )

    output = json.dumps(schema, indent=2, ensure_ascii=False)

    if args.output:
        Path(args.output).write_text(output + "\n")
        print(f"Written to {args.output} ({len(schema['tables'])} tables)")
    else:
        print(output)


if __name__ == "__main__":
    main()
